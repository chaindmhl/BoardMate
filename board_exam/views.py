from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from .models import QuestionForm, Question, AnswerKey, TestKey, Teacher, Student, Result, PracticeResult, SubjectAnalytics, TopicAnalytics, DifficultyAnalytics
from django.views import View
from django.template.loader import render_to_string
from weasyprint import HTML
import random, io, zipfile, uuid, re
import xml.etree.ElementTree as ET
from scripts.check import detect_objects, sort_objects_by_distance, group_and_sequence
from django.http import JsonResponse
import numpy as np
import cv2, time, os, json, base64, traceback
from .forms import SignUpForm
from django.contrib.auth import logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.views.decorators.csrf import csrf_protect
from django.db import IntegrityError
from django.db.models import Q
from .forms import AnswerSheetForm
from itertools import zip_longest
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.utils import timezone
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.conf import settings
import PyPDF2
import fitz
import pdfplumber, docx, string
from docx import Document
from PyPDF2 import PdfReader
from django.core.files import File
import pandas as pd
from .config import BOARD_EXAM_TOPICS, LEVELS
from django.db import models
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.db.models import Count, Q, Avg, F
from collections import defaultdict
from django.views.decorators.http import require_http_methods
import datetime
import openai
from dotenv import load_dotenv


logo_path = os.path.join(settings.BASE_DIR, 'static', 'boardmate1.png')  # full path


####################### FOR SIGNING UP ##############################

@csrf_protect
def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            # Create and save the CustomUser instance
            user = form.save(commit=False)
            user.is_active = True
            user.password = make_password(form.cleaned_data['password'])  # Hash the password
            user.save()

            # Determine role and create the related instance
            role = form.cleaned_data.get('role')
            if role == 'teacher':
                user.is_staff = True  # <-- this line makes them staff
            else:
                user.is_staff = False

            user.save()  # save the user
            if role == 'teacher':
                Teacher.objects.create(
                    user=user,
                    last_name=form.cleaned_data.get('last_name'),
                    first_name=form.cleaned_data.get('first_name'),
                    middle_name=form.cleaned_data.get('middle_name'),
                    birthdate=form.cleaned_data.get('birthdate'),
                )
            elif role == 'student':
                course = form.cleaned_data.get('course')
                Student.objects.create(
                    user=user,
                    student_id=form.cleaned_data.get('student_id'),
                    last_name=form.cleaned_data.get('last_name'),
                    first_name=form.cleaned_data.get('first_name'),
                    middle_name=form.cleaned_data.get('middle_name'),
                    birthdate=form.cleaned_data.get('birthdate'),
                    course=course,
                )
            else:
                raise ValidationError("Invalid role selected")
        messages.success(request, "Account successfully signed up!")        
        return redirect('login')
    else:
        form = SignUpForm()
    return render(request, 'signup.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None and user.is_staff == True:
                login(request, user)
                return redirect('home')  # Redirect to the home page after login
            elif user is not None and user.is_staff == False:
                login(request, user)
                return redirect('home_student')  # Redirect to the home page after login
            else:
                # Authentication failed
                messages.error(request, 'Invalid username or password.')
        else:
            # Form is not valid
            messages.error(request, 'Invalid form submission. Please try again.')
    else:
        form = AuthenticationForm()

    return render(request, 'login.html', {'form': form})
####################### FOR DASHBOARD ##############################

def is_teacher(user):
    # Assuming you have a profile field in your User model indicating the user's role
    # This function checks if the user's profile indicates they are a teacher
    return user.profile.role == 'teacher'

@login_required
def main_dashboard(request):
    if request.user.is_authenticated:
        if is_teacher(request.user):
            return redirect('home')  # Redirect to teacher dashboard
        else:
            return redirect('student_dashboard')  # Redirect to student dashboard
    else:
        return redirect('login')  # Redirect to login page if user is not authenticated

@login_required
def home(request):
    return render(request, 'home.html')  # Corrected template name

@login_required
def home_student(request):
    return render(request, 'home_student.html')  # Corrected template name

@login_required
def student_dashboard(request):
    # Logic for student's dashboard
    return render(request, 'student_dashboard.html')

def logout_view(request):
    logout(request)
    return redirect('login')  

####################### FOR ADDING QUESTION TO QUESTION BANK ##############################

def question_bank(request):
    questions = Question.objects.all()
    # Logic to fetch questions from the database and display them
    return render(request, 'question_bank.html', {'questions': questions})
  
class Add_Question(View):
    def get(self, request):
        context = {
            'BOARD_EXAMS': list(BOARD_EXAM_TOPICS.keys()),
            'BOARD_EXAM_TOPICS_JSON': json.dumps(BOARD_EXAM_TOPICS),
            'LEVELS_JSON': json.dumps(LEVELS),
        }
        return render(request, 'add_question.html', context)

    def post(self, request):
        num_questions = len([key for key in request.POST.keys() if key.startswith('question_text_')])
        questions = []

        for i in range(1, num_questions + 1):
            question_text = request.POST.get(f'question_text_{i}')
            board_exam = request.POST.get(f'board_exam_{i}')
            subject = request.POST.get(f'subject_{i}')
            topic = request.POST.get(f'topic_{i}')
            level = request.POST.get(f'level_{i}')
            choiceA = request.POST.get(f'choiceA_{i}')
            choiceB = request.POST.get(f'choiceB_{i}')
            choiceC = request.POST.get(f'choiceC_{i}')
            choiceD = request.POST.get(f'choiceD_{i}')
            choiceE = request.POST.get(f'choiceE_{i}', '')
            # correct_answer = request.POST.get(f'correct_answer_{i}')
            image = request.FILES.get(f'image_{i}', None)
            is_general = request.POST.get(f'is_general_{i}') == 'on'

            # get the letter selected for correct_answer from the form
            correct_letter = request.POST.get(f'correct_answer_{i}')
            
            # map letter to actual choice text
            choices_map = {
                'A': choiceA,
                'B': choiceB,
                'C': choiceC,
                'D': choiceD,
                'E': choiceE,
            }
            correct_answer = choices_map.get(correct_letter, '').strip()
            
            if question_text and correct_answer:
                questions.append(Question(
                    board_exam=board_exam,
                    subject=subject,
                    topic=topic,
                    level_of_difficulty=level,
                    question_text=question_text,
                    image=image,
                    choiceA=choiceA,
                    choiceB=choiceB,
                    choiceC=choiceC,
                    choiceD=choiceD,
                    choiceE=choiceE,
                    correct_answer=correct_answer,
                    is_general=is_general
                ))

        if questions:
            Question.objects.bulk_create(questions)

        return redirect('home')

    

####################### FOR CREATING EXAMINATION ##############################   

def get_random_questions(num_questions, subject):
    # Retrieve questions from the database filtered by subject
    all_questions = list(Question.objects.filter(subject=subject))

    # Check if the number of requested questions is greater than the available questions
    if num_questions > len(all_questions):
        raise ValueError("Number of requested questions exceeds the available questions for the subject.")

    # Randomly select the specified number of questions
    selected_questions = random.sample(all_questions, num_questions)

    return selected_questions



# reuse your existing config context
context = {
    'BOARD_EXAMS': list(BOARD_EXAM_TOPICS.keys()),
    'BOARD_EXAM_TOPICS_JSON': json.dumps(BOARD_EXAM_TOPICS),
    'LEVELS_JSON': json.dumps(LEVELS),
}

def generate_set_id(board_exam):
    board_exam = board_exam.lower()

    if "civil" in board_exam:
        prefix = "CE"
    elif "mechanical" in board_exam:
        prefix = "ME"
    elif "electronics" in board_exam or "ece" in board_exam:
        prefix = "ECE"
    elif "electrical" in board_exam or "ee" in board_exam:
        prefix = "EE"
    else:
        prefix = "GEN"

    # UUID shortened to 8 characters only
    return f"{prefix}_{uuid.uuid4().hex[:8]}"


def generate_test(request):
    local_context = {
        'BOARD_EXAMS': list(BOARD_EXAM_TOPICS.keys()),
        'SUBJECTS_JSON': json.dumps(BOARD_EXAM_TOPICS),
    }
   
    # Only generate new test on POST
    if request.method == 'POST':
        board_exam = request.POST.get('board_exam')
        course = request.POST.get('course')
        subject = request.POST.get('subject', '').strip()
        topic = request.POST.get('topic', '').strip()
        num_questions = int(request.POST.get('num_questions', 0))
        easy_pct = int(request.POST.get('easy_pct', 0))
        medium_pct = int(request.POST.get('medium_pct', 0))
        hard_pct = int(request.POST.get('hard_pct', 0))

        if easy_pct + medium_pct + hard_pct != 100:
            local_context['error_message'] = "Total percentage of Easy + Moderate + Hard must equal 100%"
            return render(request, 'generate_test.html', local_context)

        # --- Step 1: Build a map of all subjects per board exam ---
        exam_subject_map = {}
        for exam, subjects in BOARD_EXAM_TOPICS.items():
            exam_subject_map[exam] = [s.lower().strip() for s in subjects.keys()]

        # --- Step 2: Build a keyword-to-exams reverse map ---
        subject_to_exams = {}
        for exam, subj_list in exam_subject_map.items():
            for subj in subj_list:
                for key in subj.split(','):  # split by comma for partial match
                    key = key.strip().lower()
                    if key not in subject_to_exams:
                        subject_to_exams[key] = set()
                    subject_to_exams[key].add(exam)

        # --- Step 3: Get specific questions ---
        specific_qs = Question.objects.filter(
            board_exam=board_exam,
            subject__icontains=subject,  # partial match
            topic__icontains=topic,      # partial match
            is_general=False
        )

        # --- Step 4: Get general questions shared by subject keywords ---
        general_qs = []
        if subject:
            subject_lower = subject.lower()
            # Collect all keywords from all subjects
            keywords = [k for k in subject_lower.split(',')]
            keyword_qs = Question.objects.none()
            for kw in keywords:
                kw = kw.strip()
                # ✅ Use icontains to match any general question subject containing this keyword
                keyword_qs |= Question.objects.filter(
                    is_general=True,
                    subject__icontains=kw
                )
            general_qs = list(keyword_qs)

        # --- Combine ---
        combined_qs = list(specific_qs) + general_qs

        # --- Remove duplicates ---
        seen_ids = set()
        filtered_questions = []
        for q in combined_qs:
            if q.id not in seen_ids:
                filtered_questions.append(q)
                seen_ids.add(q.id)

        if len(filtered_questions) < num_questions:
            local_context['error_message'] = "Not enough questions available for the selected filters."
            return render(request, 'generate_test.html', local_context)

        # --- Difficulty distribution ---
        easy_count = round(num_questions * easy_pct / 100)
        medium_count = round(num_questions * medium_pct / 100)
        hard_count = num_questions - easy_count - medium_count

        easy_qs = [q for q in filtered_questions if q.level_of_difficulty == 'Easy']
        medium_qs = [q for q in filtered_questions if q.level_of_difficulty == 'Moderate']
        hard_qs = [q for q in filtered_questions if q.level_of_difficulty == 'Hard']

        if len(easy_qs) < easy_count or len(medium_qs) < medium_count or len(hard_qs) < hard_count:
            local_context['error_message'] = "Not enough questions in one of the difficulty categories."
            return render(request, 'generate_test.html', local_context)

        # --- Random selection ---
        selected_questions = (
            random.sample(easy_qs, easy_count) +
            random.sample(medium_qs, medium_count) +
            random.sample(hard_qs, hard_count)
        )
        random.shuffle(selected_questions)

        set_a_questions = random.sample(selected_questions, len(selected_questions))
        set_b_questions = random.sample(selected_questions, len(selected_questions))

        return render(request, 'generated_test.html', {
            'set_a_questions': set_a_questions,
            'set_b_questions': set_b_questions,
            'board_exam': board_exam,
            'course': course,
            'subject': subject,
            'topic': topic,
            'set_a_id': uuid.uuid4().hex,
            'set_b_id': uuid.uuid4().hex,
        })
    

    return render(request, 'generate_test.html', local_context)




####################### FOR DOWNLOADING EXAMINATION SHEET ##############################
def map_letter_text(choices_lists, correct_text_dict):
    """
    choices_lists: list of lists, e.g. [choicesA, choicesB, choicesC, ...]
    correct_text_dict: {1: "4", 2: "Blue", ...}
    
    Returns: { "1": {"letter": "A", "text": "4"}, ... }
    """
    answer_key = {}
    num_choices = len(choices_lists)
    letters = list(string.ascii_uppercase[:num_choices])  # ['A','B','C',...]
    
    for i, correct_text in correct_text_dict.items():
        # Map letters to the corresponding choice text
        choice_map = {letters[idx]: choices_lists[idx][i-1] for idx in range(num_choices)}
        correct_letter = next((l for l, t in choice_map.items() if t == correct_text), None)
        answer_key[str(i)] = {"letter": correct_letter, "text": correct_text}
    
    return answer_key

def download_test_pdf(request):
    if request.method == 'POST':
        try:
            # Retrieve the form data
            subject = request.POST.get('subject')
            board_exam = request.POST.get('board_exam')
            set_a_questions = request.POST.getlist('set_a_questions[]')
            set_a_choicesA = request.POST.getlist('set_a_choicesA[]')
            set_a_choicesB = request.POST.getlist('set_a_choicesB[]')
            set_a_choicesC = request.POST.getlist('set_a_choicesC[]')
            set_a_choicesD = request.POST.getlist('set_a_choicesD[]')
            set_a_choicesE = request.POST.getlist('set_a_choicesE[]')

            set_b_questions = request.POST.getlist('set_b_questions[]')
            set_b_choicesA = request.POST.getlist('set_b_choicesA[]')
            set_b_choicesB = request.POST.getlist('set_b_choicesB[]')
            set_b_choicesC = request.POST.getlist('set_b_choicesC[]')
            set_b_choicesD = request.POST.getlist('set_b_choicesD[]')
            set_b_choicesE = request.POST.getlist('set_b_choicesE[]')

            # Generate unique IDs for Set A and Set B
            # set_a_id = uuid.uuid4().hex
            # set_b_id = uuid.uuid4().hex
            set_a_id = generate_set_id(board_exam)
            set_b_id = generate_set_id(board_exam)


            # Retrieve correct answers for Set A
            set_a_question_ids = request.POST.getlist('set_a_question_ids[]')
            set_b_question_ids = request.POST.getlist('set_b_question_ids[]')

            set_a_correct_answers = {
                i: Question.objects.get(id=qid).correct_answer
                for i, qid in enumerate(set_a_question_ids, start=1)
            }

            set_b_correct_answers = {
                i: Question.objects.get(id=qid).correct_answer
                for i, qid in enumerate(set_b_question_ids, start=1)
            }


            # Prepare questions and image URLs for Set A
            set_a_question_data = []
            for qid, question_text in zip(set_a_question_ids, set_a_questions):
                question_obj = Question.objects.get(id=qid)
                image_a = question_obj.image if question_obj.image else None
                image_url_a = image_a.url if image_a else None
                set_a_question_data.append({
                    "question": question_text,
                    "image_url": image_url_a
                })


            # Create TestKey for Set A
            TestKey.objects.create(
                set_id=set_a_id, board_exam=board_exam, subject=subject, image=None,  # General image field can be used for overall test image
                questions=set_a_question_data,  # Store the full question data with image URLs
                choiceA=set_a_choicesA, choiceB=set_a_choicesB,
                choiceC=set_a_choicesC, choiceD=set_a_choicesD, choiceE=set_a_choicesE
            )

            # Prepare questions and image URLs for Set B
            set_b_question_data = []
            for qid, question_text in zip(set_b_question_ids, set_b_questions):
                question_obj = Question.objects.get(id=qid)
                image_b = question_obj.image if question_obj.image else None
                image_url_b = image_b.url if image_b else None
                set_b_question_data.append({
                    "question": question_text,
                    "image_url": image_url_b
                })


            # Create TestKey for Set B
            TestKey.objects.create(
                set_id=set_b_id, board_exam=board_exam, subject=subject, image=None,  # General image field can be used for overall test image
                questions=set_b_question_data,  # Store the full question data with image URLs
                choiceA=set_b_choicesA, choiceB=set_b_choicesB,
                choiceC=set_b_choicesC, choiceD=set_b_choicesD, choiceE=set_b_choicesE
            )

            # Save the answer keys and tests to the database
            set_a_answer_key = map_letter_text(
                [set_a_choicesA, set_a_choicesB, set_a_choicesC, set_a_choicesD, set_a_choicesE],
                set_a_correct_answers
            )

            set_b_answer_key = map_letter_text(
                [set_b_choicesA, set_b_choicesB, set_b_choicesC, set_b_choicesD, set_b_choicesE],
                set_b_correct_answers
            )


            AnswerKey.objects.create(set_id=set_a_id, board_exam=board_exam, subject=subject, answer_key=set_a_answer_key)
            AnswerKey.objects.create(set_id=set_b_id, board_exam=board_exam, subject=subject, answer_key=set_b_answer_key)

            # Prepare questions and choices for the PDF
            def prepare_questions(questions, ids, *choices):
                question_data = []
                for qid, question_text, *choices_set in zip(ids, questions, *choices):
                    question_obj = Question.objects.get(id=qid)
                    image_url = question_obj.image.url if question_obj.image else None
                    # Build choices as list of dicts
                    choice_list = []
                    letters = ['A', 'B', 'C', 'D', 'E']
                    for letter, choice_text in zip(letters, choices_set):
                        choice_list.append({
                            'letter': letter,
                            'text': choice_text
                        })
                    question_data.append({
                        'question': question_text,
                        'choices': choice_list,
                        'image_url': image_url
                    })
                return question_data



            questions_set_a = prepare_questions(
                set_a_questions, set_a_question_ids,
                set_a_choicesA, set_a_choicesB, set_a_choicesC, set_a_choicesD, set_a_choicesE
            )
            questions_set_b = prepare_questions(
                set_b_questions, set_b_question_ids,
                set_b_choicesA, set_b_choicesB, set_b_choicesC, set_b_choicesD, set_b_choicesE
            )


            # Render PDF templates
            context_set_a = {
                'board_exam': board_exam,
                'subject': subject,
                'questions': questions_set_a,
                'set_name': "Set A",
                'set_id': set_a_id,
                'answer_key': set_a_correct_answers,
                'logo_path': logo_path, 
            }
            context_set_b = {
                'board_exam': board_exam,
                'subject': subject,
                'questions': questions_set_b,
                'set_name': "Set B",
                'set_id': set_b_id,
                'answer_key': set_b_correct_answers,
                'logo_path': logo_path, 
            }

            html_content_set_a = render_to_string('pdf_template.html', context_set_a, request=request)
            html_content_set_b = render_to_string('pdf_template.html', context_set_b, request=request)

            # Generate PDFs
            pdf_file_set_a = HTML(string=html_content_set_a, base_url=request.build_absolute_uri('/')).write_pdf()
            pdf_file_set_b = HTML(string=html_content_set_b, base_url=request.build_absolute_uri('/')).write_pdf()

            # Create a ZIP file
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                zip_file.writestr(f"generated_test_set_a_{set_a_id}.pdf", pdf_file_set_a)
                zip_file.writestr(f"generated_test_set_b_{set_b_id}.pdf", pdf_file_set_b)

            # Return the ZIP file as a response
            response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename="generated_tests.zip"'
            return response

        except Exception as e:
            print("❌ ERROR in download_test_pdf:")
            print(traceback.format_exc())  # This prints the full error trace to the terminal
            return HttpResponse(f"An error occurred: {str(e)}", status=500)

    return HttpResponse("Invalid request method", status=405)

def download_test_interface(request):
    test_keys = TestKey.objects.all().order_by('-id')
    return render(request, 'download_test.html', {'test_keys': test_keys})


def download_existing_test_pdf(request):
    set_id = request.GET.get('set_id')
    if not set_id:
        return HttpResponse("No test selected.", status=400)

    try:
        test = TestKey.objects.get(set_id=set_id)

        questions_data = test.questions or []
        choicesA = test.choiceA or []
        choicesB = test.choiceB or []
        choicesC = test.choiceC or []
        choicesD = test.choiceD or []
        choicesE = test.choiceE or []

        # Prepare the questions in the same format your PDF template expects
        pdf_questions = []
        for idx, q in enumerate(questions_data):
            pdf_questions.append((
                q.get("question"),
                choicesA[idx] if idx < len(choicesA) else "",
                choicesB[idx] if idx < len(choicesB) else "",
                choicesC[idx] if idx < len(choicesC) else "",
                choicesD[idx] if idx < len(choicesD) else "",
                choicesE[idx] if idx < len(choicesE) else "",
                q.get("image_url")
            ))

        context = {
            'board_exam': test.board_exam,
            'subject': test.subject,
            'questions': pdf_questions,
            'set_name': "Set " + test.set_id[-1],  # optional
            'set_id': test.set_id,
            'answer_key': {},  # optional, you can include AnswerKey if needed
            'logo_path': '',    # optional logo
        }

        html_content = render_to_string('pdf_template.html', context, request=request)
        pdf_file = HTML(string=html_content, base_url=request.build_absolute_uri('/')).write_pdf()

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="test_{test.set_id}.pdf"'
        return response

    except Exception as e:
        print("❌ ERROR in download_existing_test_pdf:")
        print(traceback.format_exc())
        return HttpResponse(f"An error occurred: {str(e)}", status=500)


####################### FOR UPLOADING MOODLE XML FILE (QUESTIONS) TO THE QUESTION BANK  ##############################

def strip_tags(html):
    # Regular expression to remove HTML tags
    return re.sub('<[^<]+?>', '', html)

def extract_and_save_questions(xml_file, subject):
    # Parse the XML file
    tree = ET.parse(xml_file)
    root = tree.getroot()

    # Iterate over each question
    for question in root.findall('.//question'):
        print("Processing question...")
        # Check if 'questiontext' tag exists
        question_text_element = question.find('questiontext')
        if question_text_element is not None:
            # Extract question text and image data
            question_text, image_file = extract_question_text_and_image(question_text_element, subject)
            print('question:', question_text)
            print("Image File:", image_file)
        else:
            question_text = ''
            image_file = None
            print("No 'questiontext' found for this question.") 

        # Print image file for debugging
        print("Image File:", image_file)
        print('question:', question_text)
        # Initialize variables to store correct answer and choices
        correct_answer = ''
        choices = [''] * 5  # Initialize with empty strings

        # Map the correct answer to its corresponding letter (A to E)
        answer_letter_map = {}

        # Check if question has multiple choices
        answers = question.findall('answer')
        if len(answers) < 2:
            # Skip this question if it doesn't have multiple choices
            continue

        # Iterate over each answer
        for i, answer in enumerate(answers):
            text = strip_tags(answer.find('text').text.strip())
            fraction = int(answer.get('fraction'))
            if fraction == 100:
                # Extract the correct answer and map it to the corresponding letter
                correct_answer = chr(65 + i)  # A corresponds to 65 in ASCII
            # Map the choices to letters (A to E)
            # choices[i] = f'{chr(65 + i)}. {text}'
            choices[i] = f'{text}'

        # Skip this question if both question text and choices are empty
        if not any([question_text, any(choices)]):
            continue

        # Save the image file to the local directory and get its path
        if image_file:
            image_path = save_image_locally(image_file)
            print("Image Path:", image_path)
        else:
            image_path = None

        # Print image path for debugging
        print("Image Path:", image_path)

        # Create an instance of your Django model and save it to the database
        question_instance = Question.objects.create(
            subject=subject,
            question_text=question_text,
            image=image_path,
            choiceA=choices[0],
            choiceB=choices[1],
            choiceC=choices[2],
            choiceD=choices[3],
            choiceE=choices[4],
            correct_answer=correct_answer
        )


def extract_question_text_and_image(question_text_element, subject):
    question_text = ''
    image_files = None

    # Check if the 'text' tag exists under 'questiontext'
    text_element = question_text_element.find('./text')
    if text_element is not None:
        # Extract the text content between <p> and </p> tags, or between <p> and <img> tags if present
        text_content = text_element.text.strip()
        if text_content:
            # Extract text between <p> and <img> tags
            match = re.search(r'<p>(.*?)<img', text_content)
            if match:
                question_text = match.group(1).strip()
            else:
                # Extract text between <p> and </p> tags if <img> tag is not present
                match = re.search(r'<p>(.*?)</p>', text_content)
                if match:
                    question_text = match.group(1).strip()

        # Check if there are file tags
        file_elements = question_text_element.findall('./file')
        for file_element in file_elements:
            image_name = file_element.get('name')
            if image_name.endswith('.png') or image_name.endswith('.jpg'):
                # Decode the base64 content of the file element
                file_content_base64 = file_element.text.strip()
                file_data = base64.b64decode(file_content_base64)
                # Create an InMemoryUploadedFile object for the image
                image_file = InMemoryUploadedFile(
                    ContentFile(file_data),
                    None,
                    image_name,  # Use the file name as the image name
                    'image/jpeg',  # Assuming the image is JPEG format
                    len(file_data),
                    None
                )
    else:
        print("No 'text' tag found under 'questiontext'.")

    return question_text, image_file




def save_image_locally(image_file):
    # Get the media directory
    media_dir = os.path.join(settings.MEDIA_ROOT, 'question_images')
    # Ensure the media directory exists
    os.makedirs(media_dir, exist_ok=True)
    # Construct the filename without the '/media/' part
    _, filename = os.path.split(image_file.name)
    # Save the image file to the media directory
    image_path = os.path.join(media_dir, filename)
    with open(image_path, 'wb') as f:
        f.write(image_file.read())
    # return image_path
    return os.path.join('question_images', filename)  # Return the relative path


def upload_xml(request):
    print(request.FILES)  # Print the contents of request.FILES
    if request.method == 'POST' and 'xml_file' in request.FILES:
        xml_file = request.FILES['xml_file']
        subject = request.POST.get('subject') 
        extract_and_save_questions(xml_file, subject)
        return HttpResponse("XML file uploaded. Questions are successfully stored in Question Bank.")
    return render(request, 'upload_xml.html')

################### FOR UPLOADING DOCX, PDF, TXT ##############################

def extract_text_from_pdf(uploaded_file):
    reader = PdfReader(uploaded_file)
    text = ""
    for page in reader.pages:
        text += page.extract_text() + "\n"
    return text

def extract_text_from_docx(uploaded_file):
    doc = Document(uploaded_file)
    return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])

def extract_text_from_txt(uploaded_file):
    text = uploaded_file.read().decode('utf-8-sig')  # removes BOM automatically
    return text

def extract_and_save_questions_from_text(text, course, subject, topic, image_files=None):
    """
    Extracts questions, choices, correct answers, image references, difficulty level (E/M/H),
    and generalization (Y/N) from a text-based file.
    """

    # Normalize image filenames
    normalized_images = {}
    if image_files:
        for name, file_obj in image_files.items():
            clean_name = os.path.basename(name).strip().lower()
            normalized_images[clean_name] = file_obj

    lines = [line.strip() for line in text.strip().split("\n") if line.strip()]
    questions = []
    current_question = None

    for idx, line in enumerate(lines):
        # Detect new question
        question_match = re.match(r'^(\d+)\.\s*(.+)', line)
        if question_match:
            if current_question:
                questions.append(current_question)

            current_question = {
                "question_text": question_match.group(2).strip(),
                "choices": {"A": "", "B": "", "C": "", "D": "", "E": ""},
                "correct_answer": "",
                "image_name": None,
                "level": "Easy",
                "is_general": False,
            }
            continue

        # Detect image
        image_match = re.match(r'Image[:\-]?\s*(.+\.(?:jpg|jpeg|png|gif))', line, re.IGNORECASE)
        if image_match and current_question:
            image_name = image_match.group(1).strip().lower()
            current_question["image_name"] = image_name
            continue

        # Detect correct answer
        # correct_choice = re.match(r'>([A-E])\.\s*(.+)', line)
        # if correct_choice and current_question:
        #     letter, choice = correct_choice.groups()
        #     current_question["choices"][letter] = choice.strip()
        #     current_question["correct_answer"] = letter
        #     continue
        correct_choice = re.match(r'>([A-E])\.\s*(.+)', line)
        if correct_choice and current_question:
            letter, choice = correct_choice.groups()
            current_question["choices"][letter] = choice.strip()
            # Store the actual choice text instead of the letter
            current_question["correct_answer"] = choice.strip()
            continue


        # Detect normal choice
        normal_choice = re.match(r'([A-E])\.\s*(.+)', line)
        if normal_choice and current_question:
            letter, choice = normal_choice.groups()
            current_question["choices"][letter] = choice.strip()
            continue

        # Detect level (E/M/H)
        if current_question and line.upper() in ['E', 'M', 'H']:
            level_map = {'E': 'Easy', 'M': 'Moderate', 'H': 'Hard'}
            current_question["level"] = level_map[line.upper()]
            continue

        # Detect generalization (Y/N)
        if current_question and line.upper() in ['Y', 'N']:
            current_question["is_general"] = (line.upper() == 'Y')
            continue

        # End of file — append last question
        if idx == len(lines) - 1 and current_question:
            questions.append(current_question)

    if current_question and current_question not in questions:
        questions.append(current_question)

    # === Save to DB ===
    for q in questions:
         # Ensure choiceE always has a default
        if not q["choices"]["E"].strip():
            q["choices"]["E"] = "None of the Above"
            
        question_obj = Question.objects.create(
            board_exam=course,
            subject=subject,
            topic=topic,
            question_text=q["question_text"],
            choiceA=q["choices"]["A"],
            choiceB=q["choices"]["B"],
            choiceC=q["choices"]["C"],
            choiceD=q["choices"]["D"],
            choiceE=q["choices"]["E"],
            correct_answer=q["correct_answer"],
            level_of_difficulty=q["level"],
            is_general=q["is_general"],
        )

        # Attach image (if found)
        if q["image_name"]:
            clean_img = os.path.basename(q["image_name"]).strip().lower()
            if clean_img in normalized_images:
                file_obj = normalized_images[clean_img]
                question_obj.image.save(clean_img, ContentFile(file_obj.read()), save=True)
            else:
                print(f"⚠️ Image '{q['image_name']}' not found in uploaded files.")

def extract_and_save_from_xlsx(uploaded_file, course, subject, topic, image_map):
    df = pd.read_excel(uploaded_file)

    for _, row in df.iterrows():
        img_name = str(row.get('Image', '')).strip()
        image_file = image_map.get(img_name) if img_name else None

        choices_map = {
                'A': row.get('Choice A', '').strip(),
                'B': row.get('Choice B', '').strip(),
                'C': row.get('Choice C', '').strip(),
                'D': row.get('Choice D', '').strip(),
                'E': row.get('Choice E', '').strip() or "None of the Above",
            }
        correct_letter = str(row.get('Correct Answer', '')).upper().strip()
        correct_answer = choices_map.get(correct_letter, '')

        question = Question(
            board_exam=course,
            subject=subject,
            topic=topic,
            question_text=row.get('Question', ''),
            choiceA=row.get('Choice A', ''),
            choiceB=row.get('Choice B', ''),
            choiceC=row.get('Choice C', ''),
            choiceD=row.get('Choice D', ''),
            choiceE=row.get('Choice E', '') or "None of the Above",
            correct_answer=correct_answer,
            level_of_difficulty={'E':'Easy','M':'Moderate','H':'Hard'}.get(str(row.get('Level','')).upper(),'Easy'),
            is_general=str(row.get('Is General','')).upper()=='Y',
        )
        if image_file:
            question.image.save(image_file.name, File(image_file), save=False)
        question.save()




def upload_file(request):
    if request.method == 'POST':
        uploaded_items = request.FILES.getlist('folder_upload')
        course = request.POST.get('course')
        subject = request.POST.get('subject')
        topic = request.POST.get('topic')

        # Separate main file and image files
        main_file = None
        image_map = {}

        for f in uploaded_items:
            ext = os.path.splitext(f.name)[1].lower()
            if ext in ['.docx', '.pdf', '.txt', '.xlsx']:
                main_file = f
            elif ext in ['.jpg', '.jpeg', '.png']:
                image_map[os.path.basename(f.name)] = f

        if not main_file:
            return HttpResponse("No main file (.docx, .txt, .pdf, or .xlsx) found in the folder.")

        # Determine file type and process
        ext = os.path.splitext(main_file.name)[1].lower()
        if ext == '.pdf':
            text = extract_text_from_pdf(main_file)
            extract_and_save_questions_from_text(text, course, subject, topic, image_map)
        elif ext == '.docx':
            text = extract_text_from_docx(main_file)
            extract_and_save_questions_from_text(text, course, subject, topic, image_map)
        elif ext == '.txt':
            text = extract_text_from_txt(main_file)
            extract_and_save_questions_from_text(text, course, subject, topic, image_map)
        elif ext == '.xlsx':
            extract_and_save_from_xlsx(main_file, course, subject, topic, image_map)
        else:
            return HttpResponse("Invalid file type in folder.")

        return HttpResponse("""
            <script>
                alert("✅ Questions added successfully!");
                window.location.href = '/home/';
            </script>
        """)

    # 🧩 This is the part you were missing!
    context = {
        'BOARD_EXAMS': list(BOARD_EXAM_TOPICS.keys()),
        'BOARD_EXAM_TOPICS_JSON': json.dumps(BOARD_EXAM_TOPICS),
        'LEVELS_JSON': json.dumps(LEVELS),
    }

    return render(request, 'upload_file.html', context)





####################### FOR UPLOADING AND CHECKING OF ANSWER SHEET (IMAGE) ##############################

def get_exam_id_suggestions(request):
    input_text = request.GET.get('input', '')

    # Filter AnswerKey objects based on partial match of input_text
    suggestions = AnswerKey.objects.filter(set_id__icontains=input_text).values_list('set_id', flat=True)

    return JsonResponse(list(suggestions), safe=False)

def get_subjects(request):
    subjects = TestKey.objects.values_list('subject', flat=True).distinct()
    return JsonResponse({'subjects': list(subjects)})

def get_testkeys_by_subject(request):
    subject = request.GET.get('subject')
    testkeys = TestKey.objects.filter(subject=subject).values_list('set_id', flat=True)
    return JsonResponse({'testkeys': list(testkeys)})

def download_answer_page(request):
    return render(request, 'download_answer_key.html')

def download_exam_results_page(request):
    return render(request, 'download_exam_results.html')

def view_answer_key(request):
    exam_id = request.GET.get('exam_id')
    answer_key = get_object_or_404(AnswerKey, set_id=exam_id)

    return render(request, 'view_answer_key.html', {
        'exam_id': exam_id,
        'answer_key': answer_key.answer_key
    })

def download_answer_key(request):
    exam_id = request.GET.get('exam_id', None)

    if exam_id is None:
        return JsonResponse({'error': 'Exam ID is required'})

    answer_key = AnswerKey.objects.filter(set_id=exam_id).first()

    if answer_key is None:
        return JsonResponse({'error': 'Answer key not found for the provided exam ID'})

    # Generate file name
    file_name = f'answer_key_{exam_id}.txt'

    # Create readable text format
    answer_key_str = (
        f"Board Exam/Course: {answer_key.board_exam}\n"
        f"Subject: {answer_key.subject}\n"
        # f"Topic: {answer_key.topic}\n"
        f"Test Key: {answer_key.set_id}\n"
        f"{'-'*40}\n"
        + '\n'.join([f'{key}: {value}' for key, value in answer_key.answer_key.items()])
    )

    # Return as downloadable text file
    response = HttpResponse(answer_key_str, content_type='text/plain')
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    return response


# Get distinct board exams
def get_board_exams(request):
    board_exams = list(
        Question.objects.values_list('board_exam', flat=True).distinct()
    )
    return JsonResponse({'board_exams': board_exams})


# Get distinct subjects by board exam
def get_subjects_by_board_exam(request):
    board_exam = request.GET.get('board_exam')
    subjects = []
    if board_exam:
        subjects = list(
            Question.objects.filter(board_exam=board_exam)
            .values_list('subject', flat=True)
            .distinct()
        )
    return JsonResponse({'subjects': subjects})


# Get distinct topics by subject
def get_topics_by_subject(request):
    subject = request.GET.get('subject')
    topics = []
    if subject:
        topics = list(
            Question.objects.filter(subject=subject)
            .values_list('topic', flat=True)
            .distinct()
        )
    return JsonResponse({'topics': topics})


# Get test keys by topic (from AnswerKey)
def get_testkeys_by_topic(request):
    topic = request.GET.get('topic')
    testkeys = []
    if topic:
        # If your AnswerKey has 'subject' field, we can match it to Question.subject of this topic
        subject = Question.objects.values_list('subject', flat=True).first()
        if subject:
            testkeys = list(
                AnswerKey.objects.filter(subject=subject)
                .values_list('set_id', flat=True)
            )
    return JsonResponse({'testkeys': testkeys})

def download_exam_results(request):
    exam_id = request.GET.get('exam_id', None)

    if exam_id is None:
        return JsonResponse({'error': 'Exam ID is required'})

    # Try to get AnswerKey for metadata
    answer_key = AnswerKey.objects.filter(set_id=exam_id).first()
    if not answer_key:
        return JsonResponse({'error': 'No answer key found for this exam ID'})

    board_exam = getattr(answer_key, 'board_exam', 'N/A')
    subject = getattr(answer_key, 'subject', 'N/A')

    # Get results
    results = Result.objects.filter(exam_id=exam_id).order_by('student_name')
    if not results.exists():
        return JsonResponse({'error': 'No results found for this exam ID'})

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Exam Results for " + str(board_exam) + " (" + str(exam_id) + ")"

    # === Styles ===
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # === Header Info ===
    ws["A1"] = "Board Exam:"
    ws["B1"] = board_exam
    ws["A2"] = "Subject:"
    ws["B2"] = subject
    ws["A3"] = "Test Key:"
    ws["B3"] = exam_id

    # Blank row
    ws.append([])

    # === Table Header ===
    ws.append(["Student Name", "Score"])
    for cell in ws[5]:
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    # === Table Data ===
    for result in results:
        ws.append([result.student_name or "Unknown", result.score or "N/A"])

    # Apply border and alignment to data cells
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # === Auto column width ===
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # === Prepare response ===
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    file_name = f"{board_exam}_exam_results_{exam_id}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    wb.save(response)
    return response

# def image_to_mask(image):
#     # Convert image to grayscale
#     gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
#     # Apply adaptive thresholding
#     _, mask = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    
#     # Apply morphological operations to clean up the mask
#     kernel_size = (2, 2)
#     kernel = np.ones(kernel_size, np.uint8)
#     mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
    
#     return mask

from django_q.tasks import async_task

def upload_answer(request):
    if request.method == 'POST' and request.FILES.get('image'):
        uploaded_image = request.FILES['image']
        exam_id = request.POST.get('exam_id')
        user_id = request.user.id

        # Create uploads folder inside MEDIA_ROOT if it doesn't exist
        upload_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
        os.makedirs(upload_dir, exist_ok=True)

        # Save uploaded image with unique name
        filename = f"{int(time.time())}_{uuid.uuid4().hex}.jpg"
        file_path = os.path.join(upload_dir, filename)
        with open(file_path, 'wb') as f:
            for chunk in uploaded_image.chunks():
                f.write(chunk)

        # Relative path to pass to task
        relative_path = os.path.join("uploads", filename)

        # Queue heavy processing task
        async_task("board_exam.tasks.process_uploaded_answer", relative_path, exam_id, user_id)

        return JsonResponse({
            "status": "processing",
            "message": "Your answer is being processed."
        })

    # GET request or no file
    return render(request, "upload_answer.html")

def answer_sheet_view(request):
    if request.method == 'POST':
        form = AnswerSheetForm(request.POST)
        if form.is_valid():
            # Process the form data and save it to the database
            # Redirect to a success page or render a confirmation message
            pass  # Placeholder for processing form data
    else:
        form = AnswerSheetForm()
    return render(request, 'answer_sheet.html', {'form': form})

def online_answer_test(request):
    if request.method == 'POST':
        subject = request.POST.get('subject')
        board_exam = request.POST.get('board_exam')

        # Use the set IDs passed from generate_test.html, or generate new ones if missing
        # set_a_id = request.POST.get('set_a_id') or uuid.uuid4().hex
        # set_b_id = request.POST.get('set_b_id') or uuid.uuid4().hex
        set_a_id = generate_set_id(board_exam)
        set_b_id = generate_set_id(board_exam)


        # Get question IDs and texts
        set_a_question_ids = request.POST.getlist('set_a_question_ids[]')
        set_b_question_ids = request.POST.getlist('set_b_question_ids[]')
        set_a_questions = request.POST.getlist('set_a_questions[]')
        set_b_questions = request.POST.getlist('set_b_questions[]')

        # Get choices
        set_a_choices = [request.POST.getlist(f'set_a_choices{ch}[]') for ch in 'ABCDE']
        set_b_choices = [request.POST.getlist(f'set_b_choices{ch}[]') for ch in 'ABCDE']

        # Prepare question data and correct answers for TestKey and AnswerKey
        def prepare_question_data(question_ids, question_texts):
            data = []
            answers = {}
            for i, qid in enumerate(question_ids, start=1):
                try:
                    q = Question.objects.get(id=qid)
                    data.append({
                        "question": q.question_text,
                        "image_url": q.image.url if q.image else None
                    })
                    answers[i] = q.correct_answer
                except Question.DoesNotExist:
                    # fallback if question not found
                    data.append({
                        "question": question_texts[i-1],
                        "image_url": None
                    })
                    answers[i] = None
            return data, answers

        set_a_question_data, set_a_correct_answers = prepare_question_data(set_a_question_ids, set_a_questions)
        set_b_question_data, set_b_correct_answers = prepare_question_data(set_b_question_ids, set_b_questions)

        # Save TestKey and AnswerKey if not already existing
        if not TestKey.objects.filter(set_id=set_a_id).exists():
            TestKey.objects.create(
                set_id=set_a_id,
                board_exam=board_exam,
                subject=subject,
                questions=set_a_question_data,
                choiceA=set_a_choices[0],
                choiceB=set_a_choices[1],
                choiceC=set_a_choices[2],
                choiceD=set_a_choices[3],
                choiceE=set_a_choices[4]
            )
            set_a_answer_key = map_letter_text(set_a_choices, set_a_correct_answers)
            AnswerKey.objects.create(set_id=set_a_id, board_exam=board_exam, subject=subject, answer_key=set_a_answer_key)

        if not TestKey.objects.filter(set_id=set_b_id).exists():
            TestKey.objects.create(
                set_id=set_b_id,
                board_exam=board_exam,
                subject=subject,
                questions=set_b_question_data,
                choiceA=set_b_choices[0],
                choiceB=set_b_choices[1],
                choiceC=set_b_choices[2],
                choiceD=set_b_choices[3],
                choiceE=set_b_choices[4]
            )
            set_b_answer_key = map_letter_text(set_b_choices, set_b_correct_answers)
            AnswerKey.objects.create(set_id=set_b_id, board_exam=board_exam, subject=subject, answer_key=set_b_answer_key)

        # Prepare questions for preview
        set_a_questions_choices = list(zip(set_a_question_ids, set_a_questions, zip(*set_a_choices)))
        set_b_questions_choices = list(zip(set_b_question_ids, set_b_questions, zip(*set_b_choices)))

        # Prepare image mapping for preview
        set_a_images = {
            qid: Question.objects.get(id=qid).image.url if Question.objects.filter(id=qid).exists() and Question.objects.get(id=qid).image else None
            for qid in set_a_question_ids
        }
        set_b_images = {
            qid: Question.objects.get(id=qid).image.url if Question.objects.filter(id=qid).exists() and Question.objects.get(id=qid).image else None
            for qid in set_b_question_ids
        }

        return render(request, 'answer_test.html', {
            'subject': subject,
            'board_exam': board_exam,
            'set_a_questions_choices': set_a_questions_choices,
            'set_b_questions_choices': set_b_questions_choices,
            'set_a_images': set_a_images,
            'set_b_images': set_b_images,
            'set_a_id': set_a_id,
            'set_b_id': set_b_id,
        })
    
    return render(request, 'answer_test_form.html')


def answer_test_preview(request, subject, board_exam, set_a_id, set_b_id):
    test_a = TestKey.objects.get(set_id=set_a_id)
    test_b = TestKey.objects.get(set_id=set_b_id)

    answer_a = AnswerKey.objects.get(set_id=set_a_id)
    answer_b = AnswerKey.objects.get(set_id=set_b_id)

    return render(request, 'answer_test.html', {
        'subject': subject,
        'board_exam': board_exam,
        'test_a': test_a,
        'test_b': test_b,
        'answer_a': answer_a,
        'answer_b': answer_b,
    })


####################### FOR ANSWERING ONLINE ##############################

def answer_online_exam(request):
    if request.method == 'POST':
        set_id = request.POST.get('set_id')
        if set_id:
            return redirect('exam_form', set_id=set_id)
        messages.error(request, "Please select an exam set.")
    
    # Load board exams and their sets
    board_exams_qs = TestKey.objects.values_list('board_exam', flat=True).distinct()
    board_exams = list(board_exams_qs)

    sets_by_board_exam = {}
    for be in board_exams:
        sets = TestKey.objects.filter(board_exam=be).values('set_id', 'subject')
        sets_by_board_exam[be] = list(sets)

    sets_by_board_exam_json = json.dumps(sets_by_board_exam)

    return render(
        request,
        'answer_online_exam.html',
        {
            'board_exams': board_exams,
            'sets_by_board_exam_json': sets_by_board_exam_json,
        }
    )



def exam_form(request, set_id):
    test_key = get_object_or_404(TestKey, set_id=set_id)
    student = get_object_or_404(Student, user=request.user)

    # Check access
    if student.course.strip().lower() != test_key.board_exam.strip().lower():
        messages.error(request, "You are not allowed to access this exam.")
        return redirect('answer_online_exam')

    # Check if already submitted
    result = Result.objects.filter(user=request.user, exam_id=set_id).first()
    if not result:
        request.session.pop('form_submitted', None)
    if result and result.is_submitted:
        messages.error(request, "You have already submitted the form.")
        return redirect('warning_page')

    # Prepare questions and choices
    question_choices = []
    letters = list(string.ascii_uppercase)[:5]  # A–E

    for i, question in enumerate(test_key.questions):
        question_text = question.get('question_text') or question.get('question')
        image_url = question.get('image_url') or question.get('image') or ''

        # All choice texts
        choice_texts = [
            test_key.choiceA[i],
            test_key.choiceB[i],
            test_key.choiceC[i],
            test_key.choiceD[i],
            test_key.choiceE[i],
        ]

        random.shuffle(choice_texts)  # shuffle the texts only

        # Assign letters A–E fixed
        choices = list(zip(letters, choice_texts))
        question_choices.append((question_text, choices, image_url))

    total_items = len(question_choices)
    MAX_ITEMS = 100
    MAX_TIME_SECONDS = 4 * 60 * 60
    total_time_limit = int((total_items / MAX_ITEMS) * MAX_TIME_SECONDS)
    per_question_time_limit = total_time_limit / total_items

    if request.method == 'POST':
        if request.session.get('form_submitted'):
            messages.error(request, "You have already submitted the form.")
            return redirect('warning_page')

        # Collect answers directly from POST
        submitted_answers = []
        for i in range(total_items):
            ans = request.POST.get(f'question_{i + 1}')
            if not ans:
                messages.error(request, f"Please select an answer for question {i + 1}.")
                context = {
                    'test_key': test_key,
                    'question_choices': question_choices,
                    'total_items': total_items,
                    'total_time_limit': int(total_time_limit),
                    'per_question_time_limit': per_question_time_limit,
                    'start_time': timezone.now().isoformat(),
                }
                return render(request, 'exam_form.html', context)
            submitted_answers.append(ans)

        start_time_str = request.POST.get('start_time')
        if start_time_str:
            start_time = datetime.datetime.fromisoformat(start_time_str)
            elapsed_td = timezone.now() - start_time
            hours, remainder = divmod(int(elapsed_td.total_seconds()), 3600)
            minutes, seconds = divmod(remainder, 60)
            elapsed_time = f"{hours}hr {minutes}min {seconds}sec"
        else:
            elapsed_time = None

        request.session['form_submitted'] = True

        # Store submitted answers directly as text
        submitted_text_answers = submitted_answers


        # Calculate score
        score = 0
        answer_key = get_object_or_404(AnswerKey, set_id=set_id)
        answer_key_list = answer_key.answer_key  # this may be list of dicts

        # Build correct_text_answers for storing
        answer_key_dict = answer_key.answer_key  # probably a dict, not a list
        correct_text_answers = []

        for i, key in enumerate(sorted(answer_key_dict.keys(), key=int)):  # sort by question number
            correct_entry = answer_key_dict[key]
            correct_text = correct_entry.get('text', '')  # now we get the right text
            correct_text_answers.append(correct_text)

            # Compare submitted text vs correct text
            if submitted_text_answers[i] == correct_text:
                score += 1


        # Save result without modifying model
        try:
            result = Result.objects.create(
                user=request.user,
                student_id=student.student_id,
                course=student.course,
                student_name=str(student),
                subject=test_key.subject,
                exam_id=set_id,
                answer=submitted_text_answers,
                correct_answer=correct_text_answers,  # only texts
                score=score,
                total_items=total_items,
                is_submitted=True,
                timestamp=timezone.now(),
                elapsed_time=elapsed_time
            )

            return HttpResponseRedirect(reverse('result_page', args=[result.id]))
        except IntegrityError:
            messages.error(request, "There was an error saving your result.")

    # GET request or errors
    context = {
        'test_key': test_key,
        'question_choices': question_choices,
        'total_items': total_items,
        'total_time_limit': int(total_time_limit),
        'per_question_time_limit': per_question_time_limit,
        'start_time': timezone.now().isoformat(),
    }
    return render(request, 'exam_form.html', context)


def result_page(request, result_id):
    result = Result.objects.get(id=result_id)
    percent = 0
    if result.total_items:
        percent = round((result.score / result.total_items) * 100, 2)
    return render(request, 'result_page.html', {
        'result': result,
        'percent': percent
    })


def warning_page(request):
    home_student_url = reverse('home_student')  # Assuming 'home_student' is the name of the URL pattern for home_student.html
    return render(request, 'submit_warning.html', {'home_student_url': home_student_url})

@login_required
def view_results(request):
    # Get the results for the logged-in user only
    user_results = Result.objects.filter(user=request.user).order_by('-timestamp')
    return render(request, 'view_results.html', {'results': user_results})

def question_analytics(request):
    # Course stats
    course_stats = (
        Question.objects.values('board_exam')
        .annotate(
            total=Count('id'),
            general=Count('id', filter=Q(is_general=True)),
            non_general=Count('id', filter=Q(is_general=False))
        )
    )

    # Difficulty stats
    difficulty_stats = (
        Question.objects.values('level_of_difficulty')
        .annotate(count=Count('id'))
    )

    # Subject distribution
    subject_distribution = (
        Question.objects.values('board_exam', 'subject')
        .annotate(total_questions=Count('id'))
    )

    # Data for charts
    course_labels = [c['board_exam'] for c in course_stats]
    total_questions = [c['total'] for c in course_stats]
    general_questions = [c['general'] for c in course_stats]
    non_general_questions = [c['non_general'] for c in course_stats]

    difficulty_labels = [d['level_of_difficulty'] for d in difficulty_stats]
    difficulty_counts = [d['count'] for d in difficulty_stats]

    context = {
        "course_stats": course_stats,
        "subject_distribution": subject_distribution,
        "course_labels": json.dumps(course_labels),
        "total_questions": json.dumps(total_questions),
        "general_questions": json.dumps(general_questions),
        "non_general_questions": json.dumps(non_general_questions),
        "difficulty_labels": json.dumps(difficulty_labels),
        "difficulty_counts": json.dumps(difficulty_counts),
    }
    return render(request, "question_analytics.html", context)

def test_analytics(request):
    results = Result.objects.all()
    courses = results.values_list('course', flat=True).distinct()
    course_data = {}

    for course in courses:
        course_results = results.filter(course=course)

        # Pass/Fail
        passed_counts = course_results.filter(score__gte=0.6 * F('total_items')).count()
        failed_counts = course_results.filter(score__lt=0.6 * F('total_items')).count()

        # Average score
        avg_score = course_results.aggregate(avg=Avg('score'))['avg'] or 0

        # Per question correct/wrong
        question_stats = defaultdict(lambda: {'correct': 0, 'wrong': 0})
        for r in course_results:
            answers = r.answer or []
            correct_answers = r.correct_answer or []
            for idx, ans in enumerate(answers):
                correct_ans = correct_answers[idx] if idx < len(correct_answers) else None
                if ans == correct_ans:
                    question_stats[idx]['correct'] += 1
                else:
                    question_stats[idx]['wrong'] += 1

        question_labels = [f'Q{i+1}' for i in range(len(question_stats))]
        correct_counts = [question_stats[i]['correct'] for i in range(len(question_stats))]
        wrong_counts = [question_stats[i]['wrong'] for i in range(len(question_stats))]

        # Top students
        top_students = course_results.order_by('-score')[:10]

        # Pre-serialize JSON for template
        passed_json = json.dumps({"passed": passed_counts, "failed": failed_counts})
        avg_json = json.dumps({"avg": avg_score})
        question_json = json.dumps({"labels": question_labels, "correct": correct_counts, "wrong": wrong_counts})

        course_data[course] = {
            'passed_counts': passed_counts,
            'failed_counts': failed_counts,
            'avg_score': avg_score,
            'question_labels': question_labels,
            'correct_counts': correct_counts,
            'wrong_counts': wrong_counts,
            'top_students': top_students,
            'passed_json': passed_json,
            'avg_json': avg_json,
            'question_json': question_json
        }

    return render(request, 'test_analytics.html', {'course_data': course_data})

# ---- Practice: start ----
@require_http_methods(["GET", "POST"])
def practice_start(request):

    # 1. Get student's board exam / course
    student_course = request.user.student.course  # adjust if different field

    # 2. Load subjects for this course only
    subjects = (
        Question.objects.filter(board_exam=student_course)
        .values_list('subject', flat=True)
        .distinct()
    )

    if request.method == "POST":

        subject = request.POST.get('subject')

        try:
            num_items = int(request.POST.get("num_items") or 10)
        except ValueError:
            num_items = 10

        # Filter questions for the student’s course + chosen subject
        qs = Question.objects.filter(board_exam=student_course, subject=subject)

        if not qs.exists():
            messages.error(request, "No questions found for this subject!")
            return redirect("practice_start")

        available = qs.count()

        if num_items > available:
            messages.error(
                request,
                f"You selected {num_items} but only {available} questions exist!"
            )
            return redirect("practice_start")

        # Random selection
        questions = list(qs)
        random.shuffle(questions)
        chosen = questions[:num_items]

        payload = []
        for q in chosen:
            payload.append({
                'id': q.id,
                'text': q.question_text,
                'image_url': q.image.url if q.image else None,
                'choices': [
                    {"key": "A", "text": q.choiceA},
                    {"key": "B", "text": q.choiceB},
                    {"key": "C", "text": q.choiceC},
                    {"key": "D", "text": q.choiceD},
                    {"key": "E", "text": q.choiceE},
                ],
                "correct": q.correct_answer,
            })

        session_id = str(uuid.uuid4())
        request.session[f"practice_{session_id}"] = {
            "board_exam": student_course,
            "subject": subject,
            "questions": payload,
            "total_items": len(payload),
        }
        request.session.modified = True

        return redirect('practice_take', session_id=session_id)

    return render(request, "practice_start.html", {
        "subjects": subjects,
        "student_course": student_course,
    })



# ---- Practice: take (render questions + timer) ----
def practice_take(request, session_id):
    sess_key = f'practice_{session_id}'
    data = request.session.get(sess_key)
    if not data:
        messages.error(request, "Practice session not found or expired.")
        return redirect('practice_start')

    # pass questions without exposing 'correct' on client (we'll keep a server copy)
    letters = list(string.ascii_uppercase)  # ['A','B','C','D',...]

    questions_for_client = []
    for qi, q in enumerate(data['questions'], start=1):
        choices = q['choices'].copy()
        random.shuffle(choices)
        for idx, choice in enumerate(choices):
            choice['display_letter'] = letters[idx]
        questions_for_client.append({
            'instance_id': qi,
            'q_id': q['id'],
            'text': q['text'],
            'image_url': q['image_url'],
            'choices': choices
        })

    context = {
        'session_id': session_id,
        'board_exam': data['board_exam'],
        'questions': questions_for_client,
        'total_items': data['total_items'],
        'total_time_limit': data.get('total_time_limit'),
        'per_question_time_limit': data.get('per_question_time_limit'),
    }
    return render(request, 'practice_take.html', context)


# # ---- Practice: submit (grade & analytics) ----
# @require_http_methods(["POST"])
# def practice_submit(request, session_id):
#     sess_key = f'practice_{session_id}'
#     data = request.session.get(sess_key)
#     if not data:
#         messages.error(request, "Practice session not found or expired.")
#         return redirect('practice_start')

#     questions = data['questions']
#     total_items = len(questions)
#     results = []
#     correct_count = 0
#     total_time_elapsed = 0.0

#     # Expect POST fields: answer_{instance_id} and time_{instance_id}
#     for i, q in enumerate(questions, start=1):
#         ans = request.POST.get(f'answer_{i}')  # will be e.g. 'A' or 'B' etc.
#         correct_text = q['correct']
#         is_correct = (ans.strip() == correct_text.strip())
#         time_spent = request.POST.get(f'time_{i}', '0')
#         try:
#             time_spent = float(time_spent)
#         except Exception:
#             time_spent = 0.0

#         # server-grade using the server-stored correct answer
#         correct_key = q['correct']  # 'A'..'E'
#         is_correct = (ans == correct_key)
#         if is_correct:
#             correct_count += 1
#         total_time_elapsed += time_spent

#         results.append({
#             'index': i,
#             'q_id': q['id'],
#             'text': q['question_text'] if False else q['text'],
#             'image_url': q.get('image_url'),
#             'selected': ans,
#             'correct': correct_key,
#             'is_correct': is_correct,
#             'time_spent': time_spent,
#         })

#     score = correct_count
#     pct = (score / total_items * 100) if total_items else 0

#     # You can keep the practice results in session if you want to show them again
#     request.session[f'practice_result_{session_id}'] = {
#         'score': score,
#         'total_items': total_items,
#         'percent': pct,
#         'results': results,
#         'total_time': total_time_elapsed,
#         'board_exam': data['board_exam'],
#         'created_at': timezone.now().isoformat(),
#     }
#     request.session.modified = True

#     # inside practice_submit after grading:
#     PracticeResult.objects.create(
#         session_id=session_id,
#         user=request.user,
#         board_exam=data['board_exam'],
#         total_items=total_items,
#         score=score,
#         percent=pct,
#         total_time=total_time_elapsed,
#         answers=results
#     )

#     # Render the results page with analytics
#     return redirect('practice_result_page', session_id=session_id)
# ---- Practice: submit (grade & analytics) ----
@require_http_methods(["POST"])
def practice_submit(request, session_id):
    sess_key = f'practice_{session_id}'
    data = request.session.get(sess_key)
    if not data:
        messages.error(request, "Practice session not found or expired.")
        return redirect('practice_start')

    questions = data['questions']
    total_items = len(questions)
    results = []
    correct_count = 0
    total_time_elapsed = 0.0

    # ---- Analytics temporary containers ----
    subject_tracker = {}   # {subject: {"correct": x, "total": y, "time": seconds}}
    topic_tracker = {}     # {topic: {"correct": x, "total": y, "time": seconds}}
    difficulty_tracker = {}  # {difficulty: {"correct": x, "total": y}}
    # -----------------------------------------

    for i, q in enumerate(questions, start=1):

        ans = request.POST.get(f'answer_{i}')
        time_spent = float(request.POST.get(f'time_{i}', '0') or 0)

        correct_key = q['correct']
        is_correct = (ans == correct_key)

        # Fetch real question object
        q_obj = Question.objects.get(id=q['id'])
        subject = q_obj.subject
        topic = q_obj.topic
        difficulty = q_obj.level_of_difficulty

        # ------- Subject Tracker -------
        if subject not in subject_tracker:
            subject_tracker[subject] = {"correct": 0, "total": 0, "time": 0}
        subject_tracker[subject]["total"] += 1
        subject_tracker[subject]["time"] += time_spent
        if is_correct:
            subject_tracker[subject]["correct"] += 1

        # ------- Topic Tracker -------
        if topic not in topic_tracker:
            topic_tracker[topic] = {"correct": 0, "total": 0, "time": 0}
        topic_tracker[topic]["total"] += 1
        topic_tracker[topic]["time"] += time_spent
        if is_correct:
            topic_tracker[topic]["correct"] += 1

        # ------- Difficulty Tracker -------
        if difficulty not in difficulty_tracker:
            difficulty_tracker[difficulty] = {"correct": 0, "total": 0}
        difficulty_tracker[difficulty]["total"] += 1
        if is_correct:
            difficulty_tracker[difficulty]["correct"] += 1

        # Count correct answers
        if is_correct:
            correct_count += 1

        total_time_elapsed += time_spent

        results.append({
            'index': i,
            'q_id': q['id'],
            'text': q['text'],
            'image_url': q.get('image_url'),
            'selected': ans,
            'correct': correct_key,
            'is_correct': is_correct,
            'time_spent': time_spent,
        })

    # Final score
    score = correct_count
    pct = (score / total_items * 100) if total_items else 0

    # ---- Update SUBJECT ANALYTICS ----
    for subject, stats in subject_tracker.items():
        obj, _ = SubjectAnalytics.objects.get_or_create(
            user=request.user,
            subject=subject,
            board_exam=data['board_exam']
        )

        obj.total_items_answered += stats["total"]
        obj.total_correct += stats["correct"]
        obj.total_attempts += 1

        # Update average time
        prev_avg = obj.average_time_per_item
        new_avg = stats["time"] / stats["total"]
        obj.average_time_per_item = (prev_avg + new_avg) / 2

        obj.save()

    # ---- Update TOPIC ANALYTICS ----
    for topic, stats in topic_tracker.items():
        obj, _ = TopicAnalytics.objects.get_or_create(
            user=request.user,
            subject=data['board_exam'],  # OR q_obj.subject, depends on your design
            topic=topic
        )

        obj.total_items_answered += stats["total"]
        obj.total_correct += stats["correct"]

        prev_avg = obj.average_time_per_item
        new_avg = stats["time"] / stats["total"]
        obj.average_time_per_item = (prev_avg + new_avg) / 2

        obj.save()

    # ---- Update DIFFICULTY ANALYTICS ----
    for difficulty, stats in difficulty_tracker.items():
        obj, _ = DifficultyAnalytics.objects.get_or_create(
            user=request.user,
            board_exam=data['board_exam'],
            difficulty=difficulty
        )

        obj.total_items_answered += stats["total"]
        obj.total_correct += stats["correct"]
        obj.save()

    # ---- Save raw practice result ----
    PracticeResult.objects.create(
        session_id=session_id,
        user=request.user,
        board_exam=data['board_exam'],
        total_items=total_items,
        score=score,
        percent=pct,
        total_time=total_time_elapsed,
        answers=results
    )

    # Keep results in session
    request.session[f'practice_result_{session_id}'] = {
        'score': score,
        'total_items': total_items,
        'percent': pct,
        'results': results,
        'total_time': total_time_elapsed,
        'board_exam': data['board_exam'],
        'created_at': timezone.now().isoformat(),
    }
    request.session.modified = True

    return redirect('practice_result_page', session_id=session_id)


def practice_result_page(request, session_id):
    res = request.session.get(f'practice_result_{session_id}')
    if not res:
        messages.error(request, "No practice results found for that session.")
        return redirect('practice_start')
    return render(request, 'practice_result.html', {'res': res})

from django.core.serializers.json import DjangoJSONEncoder

@login_required
def analytics_dashboard(request):
    # Convert QuerySets to lists of dicts
    subject_data = list(
        SubjectAnalytics.objects.filter(user=request.user).values(
            'subject', 'board_exam', 'total_items_answered', 'total_correct',
            'total_attempts', 'average_time_per_item', 'last_practice_date'
        )
    )
    topic_data = list(
        TopicAnalytics.objects.filter(user=request.user).values(
            'subject', 'topic', 'total_items_answered', 'total_correct',
            'average_time_per_item', 'last_practice_date'
        )
    )
    difficulty_data = list(
        DifficultyAnalytics.objects.filter(user=request.user).values(
            'difficulty', 'total_items_answered', 'total_correct',
            'last_practice_date'
        )
    )

    # --- Prepare prompt for AI ---
    prompt = "You are an educational assistant. Here is a student's practice analytics:\n\n"

    prompt += "Subjects:\n"
    for s in subject_data:
        accuracy = round((s['total_correct'] / s['total_items_answered'] * 100), 2) if s['total_items_answered'] else 0
        prompt += f"- {s['subject']}: answered {s['total_items_answered']}, correct {s['total_correct']}, accuracy {accuracy}%, avg time {s['average_time_per_item']:.2f}s\n"

    prompt += "\nTopics:\n"
    for t in topic_data:
        accuracy = round((t['total_correct'] / t['total_items_answered'] * 100), 2) if t['total_items_answered'] else 0
        prompt += f"- {t['subject']} / {t['topic']}: answered {t['total_items_answered']}, correct {t['total_correct']}, accuracy {accuracy}%, avg time {t['average_time_per_item']:.2f}s\n"

    prompt += "\nDifficulties:\n"
    for d in difficulty_data:
        accuracy = round((d['total_correct'] / d['total_items_answered'] * 100), 2) if d['total_items_answered'] else 0
        prompt += f"- {d['difficulty']}: answered {d['total_items_answered']}, correct {d['total_correct']}, accuracy {accuracy}%\n"

    prompt += "\nBased on these analytics, suggest study habits, topics to focus on more, and strengths, in a concise manner."

    # --- Call OpenAI using new API ---
    suggestions = ""
    try:
        client = openai.OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

        response = client.chat.completions.create(
        model="gpt-3.5-turbo",  # use this for free accounts
        messages=[
            {"role": "system", "content": "You are an educational assistant."},
            {"role": "user", "content": prompt}
        ],
        temperature=0.7,
        max_tokens=300
    )

        suggestions = response.choices[0].message.content
    except Exception as e:
        suggestions = f"Could not generate suggestions: {str(e)}"

    # Pass JSON and AI suggestion to template
    return render(request, "analytics_dashboard.html", {
        "subject_analytics": json.dumps(subject_data, cls=DjangoJSONEncoder),
        "topic_analytics": json.dumps(topic_data, cls=DjangoJSONEncoder),
        "difficulty_analytics": json.dumps(difficulty_data, cls=DjangoJSONEncoder),
        "ai_suggestions": suggestions
    })
